import pandas as pd
import openpyxl

DAY_ROWS = [
    (15, 20),  # Monday
    (24, 28),  # Tuesday
    (32, 39),  # Wednesday
    (43, 50),  # Thursday
    (54, 61),  # Friday
    (66, 73),  # Weekend: Saturday & Sunday
]


def merged_size(cell, sheet):
    for rng in sheet.merged_cells:
        # print(cell.row, rng.left[0][0], ":", cell.column, rng.left[0][1])
        if cell.row == rng.left[0][0] and cell.column == rng.left[0][1]:
            return rng.size


# def getValueWithMergeLookup(sheet, cell):
#     idx = cell.coordinate
#     for range_ in sheet.merged_cell_ranges:
#         merged_cells = list(openpyxl.utils.rows_from_range(range_))
#         for row in merged_cells:
#             if idx in row:
#                 # If this is a merged cell,
#                 # return  the first cell of the merge range
#                 return sheet.cell(merged_cells[0][0]).value

#     return sheet.cell(idx).value


filename = "schedule.xlsm"
# df = pd.read_excel(
#     filename, sheet_name="Schedule", header=0, skiprows=13, usecols="A:AJ"
# )
# for col in df.columns:
#     print(col)
# print(df)

wb = openpyxl.load_workbook(filename)
sheet = wb["Schedule"]  # ["A13:AJ73"]

cell = sheet.cell(row=18, column=4)


# print(type(cell).__name__)
# print(cell.coordinate)
# print(cell.row)
# print(cell.column)
# print(cell.comment)
# print(cell.data_type)
# print(cell.parent)
# print(cell.value)

# print(sheet.merged_cells)

# for rng in sheet.merged_cells.ranges:
#     print(rng.left, rng.size)

print(merged_size(cell, sheet))
# print(sheet["A13":"AJ73"])

# row 14 is the first row of time headers

for row in sheet["B13":"AJ73"]:
    for cl in row:
        if cl.value:
            location = sheet.cell(row=cl.row, column=1).value
            start = sheet.cell(row=14, column=cl.column).value
            mrg = merged_size(cl, sheet)
            if mrg:
                end = sheet.cell(row=14, column=cl.column + mrg["columns"]).value
            else:
                end = sheet.cell(row=14, column=cl.column + 1).value
            task = cl.value
            print(
                location,
                start,
                end,
                task,
            )
